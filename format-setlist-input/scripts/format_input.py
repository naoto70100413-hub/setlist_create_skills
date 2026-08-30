#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
dance-setlist-formatter: 申込CSVを dance-setlist 入力Excelに整形するスクリプト
"""
import os, sys
os.environ.setdefault('PYTHONIOENCODING', 'utf-8')

import argparse
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Helpers ──────────────────────────────────────────────────────────────────

def read_csv_auto(path):
    """エンコーディングを自動判別して CSV を読み込む。"""
    for enc in ("utf-8-sig", "utf-8", "shift_jis", "cp932"):
        try:
            df = pd.read_csv(path, encoding=enc, dtype=str).fillna("")
            return df.apply(lambda col: col.str.strip() if col.dtype == object else col)
        except (UnicodeDecodeError, Exception):
            continue
    raise ValueError("CSVファイルの読み込みに失敗しました: {}".format(path))


def find_col(df, candidates):
    """候補列名リストから最初にマッチした列名を返す。なければ None。"""
    for c in candidates:
        if c in df.columns:
            return c
    # 部分一致フォールバック
    for c in candidates:
        matches = [col for col in df.columns if c in col]
        if matches:
            return matches[0]
    return None


def dedup_by_latest_timestamp(df, id_col, ts_col, label):
    """受付番号が重複する行がある場合、タイムスタンプが最新の行を残す。
    タイムスタンプが無い/解析できない行は最も古いものとして扱う
    （同一受付番号内で全て解析不能な場合は、元のファイル内で最後に登場した行を残す）。
    """
    if ts_col is None or ts_col not in df.columns:
        return df
    dup_ids = df.loc[df.duplicated(subset=id_col, keep=False), id_col].unique()
    if len(dup_ids) > 0:
        print("[{}] 受付番号の重複を検出: {} 件 → タイムスタンプが新しい行を採用".format(label, len(dup_ids)))
    df = df.copy()
    df["_ts_parsed"] = pd.to_datetime(df[ts_col], errors="coerce")
    # タイムスタンプ未解析(NaT)の行は最も古い扱いとし、先頭に寄せる。
    # 全て未解析の同一受付番号内では、安定ソートにより元ファイルで最後に登場した行が残る。
    df = df.sort_values("_ts_parsed", kind="stable", na_position="first")
    df = df.drop_duplicates(subset=id_col, keep="last")
    return df.drop(columns=["_ts_parsed"]).sort_index()


def normalize_entry_id(val):
    """受付番号列の書式ゆれ（「01 チーム名」等の余分な文字列付与）に対応し、
    先頭の数字部分だけを抽出する。数字が見つからなければ元の値をそのまま使う。"""
    v = str(val).strip()
    m = re.match(r"\s*(\d+)", v)
    return m.group(1) if m else v


PLACEHOLDER_TEAM_NAMES = {"", "未定", "tbd", "未確定", "調整中", "なし"}


def resolve_placeholder_name(name, entry_id):
    """出演者情報フォーム未提出等でチーム名が未定のまま複数チームがいると
    仮名が衝突するため、受付番号を付記した一意な仮名にする。"""
    n = str(name).strip()
    if n.lower() in PLACEHOLDER_TEAM_NAMES:
        return "未定(No.{})".format(entry_id)
    return n


def _cut_before_marker(val):
    """「チーム名/曲名」「チーム名（備考）」のような形式から、前半のチーム名候補を取り出す。"""
    v = val
    for marker in ("／", "/", "（", "("):
        if marker in v:
            candidate = v.split(marker)[0].strip()
            if candidate:
                v = candidate
            break
    return v


def merge_shared_teams(row, shared_cols, known_teams=None):
    """掛け持ち列群の値を収集してカンマ区切りの1文字列に統合する。
    既知のチーム名リスト(known_teams)があれば、「チーム名/曲名」形式や
    区切り文字で分割した際にチーム名自体が壊れないよう復元を試みる。"""
    known_teams = known_teams or set()
    teams = []
    for col in shared_cols:
        val = str(row.get(col, "")).strip()
        if not val:
            continue
        if val in known_teams:
            teams.append(val)
            continue
        cut_val = _cut_before_marker(val)
        if cut_val in known_teams:
            teams.append(cut_val)
            continue
        # オーバーフロー列はすでにカンマ・読点・スペース区切りの場合があるので分割。
        # ただしチーム名自体に読点等を含むケースがあるため、分割片を既知の
        # チーム名リストに対して前方から結合復元できないか試みる。
        for sep in (",", "、", "・", "　", " "):
            if sep in cut_val:
                parts = [p.strip() for p in cut_val.split(sep) if p.strip()]
                resolved, buf = [], ""
                for p in parts:
                    buf = (buf + sep + p) if buf else p
                    if buf in known_teams:
                        resolved.append(buf)
                        buf = ""
                if buf:
                    resolved.append(buf)
                teams.extend(resolved if resolved else parts)
                break
        else:
            teams.append(cut_val)
    # 重複除去（順序保持）
    seen = set()
    unique = []
    for t in teams:
        if t and t not in seen:
            seen.add(t)
            unique.append(t)
    return ", ".join(unique)


def find_duplicate_suspects(result_df):
    """曲名・アーティスト名が完全一致する行が複数の受付番号にまたがっている場合、
    重複申込の疑いとして検出する（氏名・メールは本スクリプトの入力に含まれないため、
    曲・アーティストの一致のみで判定する簡易版）。"""
    groups = {}
    for _, row in result_df.iterrows():
        key = (row["曲名"], row["アーティスト名"])
        if not key[0] and not key[1]:
            continue
        groups.setdefault(key, []).append(row["受付番号"])
    return {k: v for k, v in groups.items() if len(v) > 1}


# ── Excel Output ─────────────────────────────────────────────────────────────

HDR_FILL = PatternFill("solid", start_color="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=11)
NOTE_FILL = PatternFill("solid", start_color="E8F4FD")
NOTE_FONT = Font(color="1A5276", italic=True, name="Arial", size=9)
ALT_FILL  = PatternFill("solid", start_color="F5F5F5")
THIN = Side(style="thin", color="BBBBBB")

OUTPUT_COLS = ["受付番号", "チーム名", "曲名", "アーティスト名", "掛け持ちチーム", "希望時間帯"]
COL_WIDTHS  = [12, 18, 24, 20, 32, 14]
NOTE_TEXTS  = [
    "",
    "※ 必須",
    "※ 必須（Spotify検索）",
    "※ 必須（Spotify検索）",
    "掛け持ちメンバーがいる場合\n複数はカンマ区切り",
    "前半 / 後半 / 指定なし",
]


def bdr():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def neutralize_formula_cell(cell):
    """アーティスト名等が =LOVE のように =+-@ で始まると、Excel/openpyxlが
    数式と誤解釈して値が消えることがある。該当セルは明示的に文字列型にする。"""
    v = cell.value
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        cell.data_type = "s"


def write_excel(output_path, result_df, missing_ids, duplicate_suspects=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "出演チームリスト"

    # Header row
    for col, h in enumerate(OUTPUT_COLS, 1):
        c = ws.cell(1, col, h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr()
    ws.row_dimensions[1].height = 20

    # Data rows
    for i, row in enumerate(result_df.itertuples(index=False), 2):
        values = [row.受付番号, row.チーム名, row.曲名, row.アーティスト名, row.掛け持ちチーム, row.希望時間帯]
        for col, val in enumerate(values, 1):
            c = ws.cell(i, col, val)
            if i % 2 == 0:
                c.fill = ALT_FILL
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(
                horizontal="left" if col in (2, 3, 4, 5) else "center",
                vertical="center"
            )
            c.border = bdr()
            neutralize_formula_cell(c)
        ws.row_dimensions[i].height = 16

    # Notes row
    note_row = ws.max_row + 2
    for col, note in enumerate(NOTE_TEXTS, 1):
        c = ws.cell(note_row, col, note)
        c.fill = NOTE_FILL
        c.font = NOTE_FONT
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = bdr()
    ws.row_dimensions[note_row].height = 48

    # Warning sheet for unmatched IDs
    if missing_ids:
        ws2 = wb.create_sheet("結合エラー")
        ws2.cell(1, 1, "結合できなかった受付番号")
        ws2.cell(1, 1).font = Font(bold=True, name="Arial")
        for i, mid in enumerate(missing_ids, 2):
            ws2.cell(i, 1, mid)
        ws2.column_dimensions["A"].width = 30

    # Warning sheet for suspected duplicate submissions
    if duplicate_suspects:
        ws3 = wb.create_sheet("重複疑いチェック")
        hdrs = ["曲名", "アーティスト名", "該当受付番号", "備考"]
        ws3.append(hdrs)
        for col in range(1, len(hdrs) + 1):
            c = ws3.cell(1, col)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center"); c.border = bdr()
        for (song, artist), ids in duplicate_suspects.items():
            row = [song, artist, ", ".join(str(i) for i in ids),
                   "曲・アーティストが一致（氏名/メールは未照合）。出演者情報フォーム提出済みの方を正としてください"]
            ws3.append(row)
            ri = ws3.max_row
            for col in range(1, len(hdrs) + 1):
                c = ws3.cell(ri, col)
                c.fill = NOTE_FILL
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                c.border = bdr()
                neutralize_formula_cell(c)
        widths = [24, 20, 24, 40]
        for i, w in enumerate(widths, 1):
            ws3.column_dimensions[get_column_letter(i)].width = w

    # Column widths (main sheet)
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description="申込CSV → dance-setlist 入力Excel 整形スクリプト")
    p.add_argument("--songs",  required=True,  help="曲情報CSVのパス（受付番号・曲名・アーティスト名）")
    p.add_argument("--teams",  required=True,  help="チーム情報CSVのパス（受付番号・チーム名・掛け持ち列・希望時間帯）")
    p.add_argument("--output", default="",     help="出力Excelのパス（省略時は自動命名）")
    # Column name overrides
    p.add_argument("--col-id",            default="受付番号")
    p.add_argument("--col-song",          default="曲名")
    p.add_argument("--col-artist",        default="アーティスト名")
    p.add_argument("--col-team",          default="チーム名")
    p.add_argument("--col-pref",          default="希望時間帯")
    p.add_argument("--col-shared-prefix", default="掛け持ち",
                   help="掛け持ち列のプレフィックス（デフォルト: 掛け持ち）")
    p.add_argument("--col-timestamp", default="タイムスタンプ",
                   help="タイムスタンプ列の名前（デフォルト: タイムスタンプ）")
    args = p.parse_args()

    # --- Load ---
    df_songs = read_csv_auto(args.songs)
    df_teams = read_csv_auto(args.teams)

    # --- Detect key columns ---
    id_col_s = find_col(df_songs, [args.col_id, "受付番号", "受付No", "No", "番号", "ID"])
    id_col_t = find_col(df_teams, [args.col_id, "受付番号", "受付No", "No", "番号", "ID"])
    song_col   = find_col(df_songs, [args.col_song,   "曲名", "楽曲名", "曲"])
    artist_col = find_col(df_songs, [args.col_artist, "アーティスト名", "アーティスト", "歌手"])
    team_col   = find_col(df_teams, [args.col_team,   "チーム名", "チーム", "団体名"])
    pref_col   = find_col(df_teams, [args.col_pref,   "希望時間帯", "希望", "時間帯"])
    ts_col_s   = find_col(df_songs, [args.col_timestamp, "タイムスタンプ", "Timestamp"])
    ts_col_t   = find_col(df_teams, [args.col_timestamp, "タイムスタンプ", "Timestamp"])

    for name, val in [("受付番号(曲情報)", id_col_s), ("受付番号(チーム情報)", id_col_t),
                       ("曲名", song_col), ("アーティスト名", artist_col), ("チーム名", team_col)]:
        if val is None:
            sys.exit("[ERROR] 列が見つかりません: {}".format(name))

    # --- 受付番号の書式ゆれ（「01 チーム名」等）を吸収し、先頭の数字だけに正規化 ---
    df_songs[id_col_s] = df_songs[id_col_s].apply(normalize_entry_id)
    df_teams[id_col_t] = df_teams[id_col_t].apply(normalize_entry_id)

    # --- 受付番号の重複行はタイムスタンプが新しい方を採用 ---
    df_songs = dedup_by_latest_timestamp(df_songs, id_col_s, ts_col_s, "曲情報")
    df_teams = dedup_by_latest_timestamp(df_teams, id_col_t, ts_col_t, "チーム情報")

    # --- Detect shared-team columns ---
    shared_cols = [col for col in df_teams.columns if args.col_shared_prefix in col]
    if shared_cols:
        print("掛け持ち列を検出: {}".format(shared_cols))
    else:
        print("掛け持ち列なし（掛け持ちチーム列は空欄で出力します）")

    # 掛け持ち列の名寄せで使う既知のチーム名一覧（リネーム前の生の値）
    known_teams = {str(v).strip() for v in df_teams[team_col] if str(v).strip()}

    # --- Rename to canonical names ---
    df_songs = df_songs.rename(columns={id_col_s: "__id__", song_col: "曲名", artist_col: "アーティスト名"})
    df_teams = df_teams.rename(columns={id_col_t: "__id__", team_col: "チーム名"})
    if pref_col:
        df_teams = df_teams.rename(columns={pref_col: "希望時間帯"})
    else:
        df_teams["希望時間帯"] = ""

    # --- チーム名未定の申込は受付番号を付記した一意な仮名にする ---
    df_teams["チーム名"] = [resolve_placeholder_name(n, i)
                           for n, i in zip(df_teams["チーム名"], df_teams["__id__"])]

    # --- Merge shared team columns ---
    df_teams["掛け持ちチーム"] = df_teams.apply(
        lambda row: merge_shared_teams(row, shared_cols, known_teams), axis=1
    )

    # --- Join on __id__ ---
    songs_slim = df_songs[["__id__", "曲名", "アーティスト名"]]
    teams_slim = df_teams[["__id__", "チーム名", "掛け持ちチーム", "希望時間帯"]]

    merged = pd.merge(songs_slim, teams_slim, on="__id__", how="inner")

    # Detect unmatched
    all_ids = set(df_songs["__id__"]) | set(df_teams["__id__"])
    matched_ids = set(merged["__id__"])
    missing_ids = sorted(all_ids - matched_ids)

    # Final column order
    result = merged.rename(columns={"__id__": "受付番号"})
    result = result[["受付番号", "チーム名", "曲名", "アーティスト名", "掛け持ちチーム", "希望時間帯"]].copy()
    result = result.reset_index(drop=True)

    # --- Output ---
    if args.output:
        out_path = args.output
    else:
        out_path = "setlist_input_{}.xlsx".format(datetime.now().strftime("%Y%m%d"))

    duplicate_suspects = find_duplicate_suspects(result)
    write_excel(out_path, result, missing_ids, duplicate_suspects)

    print("FORMATTER COMPLETE: {}".format(out_path))
    print("  {} チームを出力しました".format(len(result)))
    if missing_ids:
        print("WARNING: 結合できなかった受付番号 {} 件 → 「結合エラー」シートを確認してください".format(len(missing_ids)))
        for mid in missing_ids[:5]:
            print("  - {}".format(mid))
        if len(missing_ids) > 5:
            print("  ... 他 {} 件".format(len(missing_ids) - 5))
    if duplicate_suspects:
        print("WARNING: 重複申込の疑いが {} 組 → 「重複疑いチェック」シートを確認してください".format(
            len(duplicate_suspects)))
    print("次のステップ: このファイルを dance-setlist スキルの入力として使用できます")


if __name__ == "__main__":
    main()
