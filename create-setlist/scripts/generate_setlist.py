#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""dance-setlist: ダンスイベントセットリスト生成スクリプト"""
import os, sys
os.environ.setdefault('PYTHONIOENCODING', 'utf-8')
import argparse, json, math, random, time
try:
    from ortools.sat.python import cp_model
    HAS_ORTOOLS = True
except ImportError:
    HAS_ORTOOLS = False
from datetime import datetime, timedelta
from pathlib import Path
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Spotify API ──────────────────────────────────────────────────────────────

def load_spotify_credentials():
    p = Path.home() / ".claude" / "spotify_credentials.json"
    if not p.exists(): return None, None
    c = json.load(open(p))
    return c.get("client_id"), c.get("client_secret")

def get_spotify_token(client_id, client_secret):
    r = requests.post("https://accounts.spotify.com/api/token",
        data={"grant_type": "client_credentials"}, auth=(client_id, client_secret), timeout=10)
    r.raise_for_status()
    return r.json()["access_token"]

def search_track_duration(token, track_name, artist_name, retries=3):
    q = "track:{} artist:{}".format(track_name, artist_name)
    for _ in range(retries):
        try:
            r = requests.get("https://api.spotify.com/v1/search",
                headers={"Authorization": "Bearer {}".format(token)},
                params={"q": q, "type": "track", "limit": 1}, timeout=10)
            if r.status_code == 429: time.sleep(1); continue
            r.raise_for_status()
            items = r.json().get("tracks", {}).get("items", [])
            return items[0]["duration_ms"] // 1000 if items else None
        except Exception: time.sleep(1)
    return None

# ── Scheduling ───────────────────────────────────────────────────────────────

def parse_shared(val):
    if not val or (isinstance(val, float) and math.isnan(val)): return []
    return [t.strip() for t in str(val).split(",") if t.strip()]

def strip_trailing_notes(df):
    """format-setlist-input が出力するExcelは、データ行の後に空行1行を挟んで
    注記行（※必須 等）が続く。全列が空の最初の行以降（注記行含む）を切り捨てる。"""
    def is_blank(row):
        return all(pd.isna(v) or str(v).strip() == "" for v in row)
    for pos in range(len(df)):
        if is_blank(df.iloc[pos]):
            return df.iloc[:pos].reset_index(drop=True)
    return df.reset_index(drop=True)

def pref_half(val):
    if not val or (isinstance(val, float) and math.isnan(val)): return None
    return "first" if "前半" in str(val) else ("second" if "後半" in str(val) else None)

# 希望時間帯の種別:
#   first/second = 前半/後半（従来どおり）
#   early/late   = できるだけ早め/遅めがいい（連続的な希望）
#   before/after = 指定時刻(HH:MM)より前/後に出演したい（"time"キーが必須）
PREF_TYPES = ("first", "second", "early", "late", "before", "after")

def normalize_pref(override, raw_val):
    """--pref-overrides で渡された構造化データを優先し、なければ従来通り
    希望時間帯列の生テキストから前半/後半を簡易判定する。"""
    if isinstance(override, dict) and override.get("type") in PREF_TYPES:
        t = override["type"]
        if t in ("before", "after"):
            if not override.get("time"):
                return None
            return {"type": t, "time": override["time"]}
        return {"type": t}
    half = pref_half(raw_val)
    return {"type": half} if half else None

def build_conflicts(teams):
    names = {t["name"] for t in teams}
    pairs = set()
    for t in teams:
        for o in t["shared_teams"]:
            if o in names and o != t["name"]:
                pairs.add(frozenset([t["name"], o]))
    return pairs

def build_similarity_conflicts(teams, key):
    """曲名/アーティスト名が同じチーム同士のペアを列挙する。"""
    groups = {}
    for t in teams:
        val = str(t.get(key) or "").strip()
        if not val:
            continue
        groups.setdefault(val, []).append(t["name"])
    pairs = set()
    for names in groups.values():
        for i in range(len(names)):
            for j in range(i + 1, len(names)):
                pairs.add(frozenset([names[i], names[j]]))
    return pairs

def position_gap(order, a, b):
    """a と b の間に挟まるチーム数を返す。"""
    ia, ib = order.index(a), order.index(b)
    return abs(ib - ia) - 1

def compute_block_boundaries(n, n_blocks):
    """assign_blocks と同じ分割ルールで、休憩が入る位置（0-indexed、
    その位置の直前で休憩が発生する）のリストを返す。"""
    if n_blocks <= 1 or n <= 0:
        return []
    base, extra = divmod(n, n_blocks)
    boundaries, idx = [], 0
    for i in range(n_blocks - 1):
        idx += base + (1 if i < extra else 0)
        if idx >= n:
            break
        boundaries.append(idx)
    return boundaries

# 優先度順: 1. 掛け持ち間隔 > 2. 曲・アーティストの連続回避 > 3. 希望時間帯
IDEAL_SHARED_GAP_SEC = 3600   # 基本目標: 1時間
MIN_SHARED_GAP_SEC   = 2400   # 最低ライン: 40分
SHARED_WEIGHT         = 200
MIN_SIMILARITY_GAP    = 2     # 曲・アーティストは最低2組空ける
SIMILARITY_WEIGHT     = 40
PREF_WEIGHT           = 5
PREF_DEADLINE_DECAY_SEC = 3600   # 指定時刻(before/after)を外れた場合、1時間ズレで実質0点まで減衰

def elapsed_gap(ordered, a, b, durations, boundaries=(), break_sec=0):
    ia, ib = ordered.index(a), ordered.index(b)
    if ia > ib: ia, ib = ib, ia
    gap = sum(durations.get(ordered[i], 0) + 30 for i in range(ia, ib))
    # a-b の間にブロックの休憩が挟まる場合、その分の時間も間隔に加算する
    # （休憩は最低13分は確保される前提で見積もる）
    crossed = sum(1 for p in boundaries if ia < p <= ib)
    return gap + crossed * break_sec

def estimate_start_dt(order, i, durations, boundaries, break_sec, start_dt):
    """出演順 order の position i にいるチームの、開始時刻の見積もりを返す。"""
    if start_dt is None:
        return None
    cum = sum(durations.get(order[j], 0) + 30 for j in range(i))
    cum += break_sec * sum(1 for p in boundaries if p <= i)
    return start_dt + timedelta(seconds=cum)

def pref_contribution(pref, i, n, order, durations, boundaries, break_sec, start_dt):
    if not pref:
        return 0
    t = pref["type"]
    if t == "first":
        return PREF_WEIGHT if i < n / 2 else 0
    if t == "second":
        return PREF_WEIGHT if i >= n / 2 else 0
    if t == "early":
        # 出演順が早いほど高スコア（連続的な希望）
        return PREF_WEIGHT * (1 - i / max(n - 1, 1))
    if t == "late":
        return PREF_WEIGHT * (i / max(n - 1, 1))
    if t in ("before", "after"):
        est = estimate_start_dt(order, i, durations, boundaries, break_sec, start_dt)
        if est is None:
            return 0
        try:
            deadline = datetime.strptime(pref["time"], "%H:%M").replace(
                year=est.year, month=est.month, day=est.day)
        except (ValueError, KeyError):
            return 0
        # 満たしていれば満点（余裕があっても加点しない）。外れている場合は
        # 反比例で減衰させ、ズレがどれだけ大きくても常に「近い方が高スコア」を維持する
        # （PREF_DEADLINE_DECAY_SEC ズレると満点の半分になる）。
        if t == "before":
            miss = max(0.0, (est - deadline).total_seconds())
        else:
            miss = max(0.0, (deadline - est).total_seconds())
        return PREF_WEIGHT / (1 + miss / PREF_DEADLINE_DECAY_SEC)
    return 0

def score_order(order, durations, conflicts, prefs, n, n_blocks=1, break_sec=0,
                 similarity_pairs=(), start_dt=None):
    s = 0
    boundaries = compute_block_boundaries(len(order), n_blocks)
    for pair in conflicts:
        a, b = list(pair)
        if a in order and b in order:
            g = elapsed_gap(order, a, b, durations, boundaries, break_sec)
            # 凹関数(√)でギャップを評価する。
            # 凹関数はジェンセンの不等式により、同じ合計ギャップでも
            # 均等分散のほうが常に高スコアになる。
            # 例: A-B=10分・B-C=60分 → スコア合計 281.7
            #     A-B=35分・B-C=35分 → スコア合計 305.6 ← 均等ケースが勝つ
            s += (min(g, IDEAL_SHARED_GAP_SEC) / IDEAL_SHARED_GAP_SEC) ** 0.5 * SHARED_WEIGHT
    for pairs in similarity_pairs:
        for pair in pairs:
            a, b = list(pair)
            if a in order and b in order:
                gap = position_gap(order, a, b)
                s += min(gap, MIN_SIMILARITY_GAP) / MIN_SIMILARITY_GAP * SIMILARITY_WEIGHT
    for i, name in enumerate(order):
        s += pref_contribution(prefs.get(name), i, n, order, durations, boundaries, break_sec, start_dt)
    return s

def greedy_schedule_heuristic(teams, durations, conflicts, n_blocks=1, break_sec=0,
                               song_pairs=frozenset(), artist_pairs=frozenset(),
                               pref_overrides=None, start_dt=None, randomize=False):
    """貪欲法（近似解）。ortools が使えない環境向けのフォールバック。"""
    names = [t["name"] for t in teams]
    if randomize:
        # --seed 指定時: 処理順の同点タイブレークをランダム化し、
        # 複数回試して比較できるようにする（貪欲法の局所最適回避用）
        random.shuffle(names)
    pref_overrides = pref_overrides or {}
    prefs = {t["name"]: normalize_pref(pref_overrides.get(t["name"]), t.get("preferred_time")) for t in teams}
    similarity_pairs = (song_pairs, artist_pairs)
    # 処理順は制約の優先度を反映: 掛け持ちを重く、曲・アーティスト重複を軽く重み付け
    cc = {n: 0 for n in names}
    for pair in conflicts:
        for n in pair: cc[n] += 3
    for pairs in similarity_pairs:
        for pair in pairs:
            for n in pair: cc[n] += 1
    remaining = sorted(names, key=lambda x: -cc[x])
    ordered = []
    for name in remaining:
        best_pos, best_score = 0, -float("inf")
        for pos in range(len(ordered) + 1):
            cand = ordered[:pos] + [name] + ordered[pos:]
            sc = score_order(cand, durations, conflicts, prefs, len(names), n_blocks, break_sec,
                              similarity_pairs, start_dt)
            if sc > best_score: best_score, best_pos = sc, pos
        ordered.insert(best_pos, name)
    return ordered

# ── CP-SAT (OR-Tools) による厳密探索 ───────────────────────────────────────────
#
# 貪欲法は「一度確定した配置に後戻りしない」ため局所最適に陥り得る。CP-SATは
# 組み合わせを網羅的に探索する制約ソルバーで、時間内であれば最適解（または
# それに近い解、探索過程で目的関数の上下界が分かる）を返せる。
#
# モデル化の要点:
#   - team_at_rank[r] / rank_of_team[i]: 出演順(rank)とチームの相互変換（順列）
#   - dur_at_rank[r]:  rank r にいるチームの所要時間（AddElementで参照）
#   - start_at_rank[r]: rank r の開始時刻（秒。累積和で計算、ブロック休憩は
#     compute_block_boundaries と同じ位置で break_sec を加算する近似値）
#   - start_of_team[i]: チーム i の開始時刻（rank_of_team 経由でElement参照）
# これにより「掛け持ちペアの間隔」「曲/アーティストの順位差」「希望時間帯との
# 距離」を全てteam_at_rank/rank_of_teamの整数変数から一意に計算できる。

def _cpsat_solve(teams, durations, conflicts, n_blocks, break_sec,
                  song_pairs, artist_pairs, pref_overrides, start_dt,
                  time_limit_sec=15, num_workers=8):
    names = [t["name"] for t in teams]
    n = len(names)
    if n == 0:
        return []
    if n == 1:
        return names[:]
    idx = {name: i for i, name in enumerate(names)}
    dur_list = [int(durations.get(name, 0)) for name in names]
    boundaries = set(compute_block_boundaries(n, n_blocks))
    pref_overrides = pref_overrides or {}
    prefs = {name: normalize_pref(pref_overrides.get(name), t.get("preferred_time"))
             for name, t in zip(names, teams)}

    model = cp_model.CpModel()

    team_at_rank = [model.NewIntVar(0, n - 1, "tar_{}".format(r)) for r in range(n)]
    rank_of_team = [model.NewIntVar(0, n - 1, "rot_{}".format(i)) for i in range(n)]
    model.AddInverse(team_at_rank, rank_of_team)

    max_dur = max(dur_list) if dur_list else 0
    dur_at_rank = [model.NewIntVar(0, max_dur, "dar_{}".format(r)) for r in range(n)]
    for r in range(n):
        model.AddElement(team_at_rank[r], dur_list, dur_at_rank[r])

    horizon = sum(dur_list) + 30 * n + break_sec * len(boundaries) + 10
    start_at_rank = [model.NewIntVar(0, horizon, "sar_{}".format(r)) for r in range(n)]
    model.Add(start_at_rank[0] == 0)
    for r in range(1, n):
        extra = break_sec if r in boundaries else 0
        model.Add(start_at_rank[r] == start_at_rank[r - 1] + dur_at_rank[r - 1] + 30 + extra)

    start_of_team = [model.NewIntVar(0, horizon, "sot_{}".format(i)) for i in range(n)]
    for i in range(n):
        model.AddElement(rank_of_team[i], start_at_rank, start_of_team[i])

    objective_terms = []

    # 1. 掛け持ち間隔（最優先）: 40分未満は強く回避しつつ、1時間到達を目指す。
    #    さらに「一番厳しいペア」を直接押し上げる項を加え、片方だけ1時間・
    #    もう片方が極端に短くなる不均衡を避ける。
    gap_vars = []
    for pair in conflicts:
        a, b = list(pair)
        ia, ib = idx[a], idx[b]
        diff = model.NewIntVar(-horizon, horizon, "diff_{}_{}".format(a, b))
        model.Add(diff == start_of_team[ia] - start_of_team[ib])
        gap = model.NewIntVar(0, horizon, "gap_{}_{}".format(a, b))
        model.AddAbsEquality(gap, diff)
        gap_vars.append(gap)
        capped = model.NewIntVar(0, IDEAL_SHARED_GAP_SEC, "capped_{}_{}".format(a, b))
        model.Add(capped <= gap)
        model.Add(capped <= IDEAL_SHARED_GAP_SEC)
        objective_terms.append((SHARED_WEIGHT, capped))
    if gap_vars:
        min_gap = model.NewIntVar(0, horizon, "min_shared_gap")
        model.AddMinEquality(min_gap, gap_vars)
        # 最も厳しいペアの底上げを重視（不均衡回避）。40分に達すれば頭打ち。
        min_gap_capped = model.NewIntVar(0, MIN_SHARED_GAP_SEC, "min_shared_gap_capped")
        model.Add(min_gap_capped <= min_gap)
        model.Add(min_gap_capped <= MIN_SHARED_GAP_SEC)
        objective_terms.append((SHARED_WEIGHT * 3, min_gap_capped))

    # 2. 曲・アーティストの連続回避（2番目）: 順位差(-1)を最低2組まで評価
    for pairs in (song_pairs, artist_pairs):
        for pair in pairs:
            a, b = list(pair)
            ia, ib = idx[a], idx[b]
            rdiff = model.NewIntVar(-n, n, "rdiff_{}_{}".format(a, b))
            model.Add(rdiff == rank_of_team[ia] - rank_of_team[ib])
            rabs = model.NewIntVar(0, n, "rabs_{}_{}".format(a, b))
            model.AddAbsEquality(rabs, rdiff)
            pos_gap = model.NewIntVar(0, n, "posgap_{}_{}".format(a, b))
            model.Add(pos_gap == rabs - 1)
            capped_sim = model.NewIntVar(0, MIN_SIMILARITY_GAP, "cappedsim_{}_{}".format(a, b))
            model.Add(capped_sim <= pos_gap)
            model.Add(capped_sim <= MIN_SIMILARITY_GAP)
            objective_terms.append((SIMILARITY_WEIGHT // MIN_SIMILARITY_GAP, capped_sim))

    # 3. 希望時間帯（最下位）
    half = n // 2
    for i, name in enumerate(names):
        pref = prefs.get(name)
        if not pref:
            continue
        t = pref["type"]
        if t == "first":
            b = model.NewBoolVar("pref_first_{}".format(i))
            model.Add(rank_of_team[i] < half).OnlyEnforceIf(b)
            model.Add(rank_of_team[i] >= half).OnlyEnforceIf(b.Not())
            objective_terms.append((PREF_WEIGHT, b))
        elif t == "second":
            b = model.NewBoolVar("pref_second_{}".format(i))
            model.Add(rank_of_team[i] >= half).OnlyEnforceIf(b)
            model.Add(rank_of_team[i] < half).OnlyEnforceIf(b.Not())
            objective_terms.append((PREF_WEIGHT, b))
        elif t == "early":
            # rank が小さいほど良い → -rank を最大化（重みは元の連続評価と概ね揃える）
            objective_terms.append((-1, rank_of_team[i]))
        elif t == "late":
            objective_terms.append((1, rank_of_team[i]))
        elif t in ("before", "after") and start_dt is not None:
            try:
                deadline_dt = datetime.strptime(pref["time"], "%H:%M").replace(
                    year=start_dt.year, month=start_dt.month, day=start_dt.day)
            except (ValueError, KeyError):
                continue
            deadline_sec = int((deadline_dt - start_dt).total_seconds())
            if deadline_sec < 0:
                continue
            # 未達の場合は線形にペナルティ（際限なく差を評価し続ける。
            # CP-SATは厳密探索のため、貪欲法のような「反比例で減衰」の
            # 工夫がなくても、際限のない線形差で十分「近い方を優先」できる）
            penalty_rate = PREF_WEIGHT / (2.0 * PREF_DEADLINE_DECAY_SEC)
            # over/short は「開始時刻 - 指定時刻」の差なので、horizon（総所要時間）
            # だけでなく deadline_sec 自体も超え得る。両方をカバーする上限にする。
            deadline_slack = horizon + deadline_sec + 10
            if t == "before":
                over = model.NewIntVar(0, deadline_slack, "over_{}".format(i))
                model.AddMaxEquality(over, [start_of_team[i] - deadline_sec, 0])
                objective_terms.append((-penalty_rate, over))
            else:
                short = model.NewIntVar(0, deadline_slack, "short_{}".format(i))
                model.AddMaxEquality(short, [deadline_sec - start_of_team[i], 0])
                objective_terms.append((-penalty_rate, short))

    # 目的関数の重みには小数(penalty_rate)が混在するため、全体を整数スケールに統一する。
    OBJ_SCALE = 10000
    model.Maximize(sum(int(round(w * OBJ_SCALE)) * v for w, v in objective_terms))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = time_limit_sec
    solver.parameters.num_search_workers = num_workers
    status = solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return None
    order_by_rank = [None] * n
    for r in range(n):
        order_by_rank[r] = names[solver.Value(team_at_rank[r])]
    return order_by_rank

def greedy_schedule(teams, durations, conflicts, n_blocks=1, break_sec=0,
                     song_pairs=frozenset(), artist_pairs=frozenset(),
                     pref_overrides=None, start_dt=None, randomize=False):
    """出演順を決定する。ortools(CP-SAT)が使えれば厳密探索を優先し、
    使えない/失敗した場合は貪欲法(greedy_schedule_heuristic)にフォールバックする。"""
    if HAS_ORTOOLS:
        try:
            result = _cpsat_solve(teams, durations, conflicts, n_blocks, break_sec,
                                   song_pairs, artist_pairs, pref_overrides, start_dt)
            if result is not None:
                return result
            sys.stderr.write("[CP-SAT] 実行可能解が見つからなかったため貪欲法にフォールバックします\n")
        except Exception as e:
            sys.stderr.write("[CP-SAT Error] {} → 貪欲法にフォールバックします\n".format(e))
    else:
        sys.stderr.write("[Info] ortools が未インストールのため貪欲法（近似解）を使用します\n")
    return greedy_schedule_heuristic(teams, durations, conflicts, n_blocks, break_sec,
                                      song_pairs, artist_pairs, pref_overrides, start_dt, randomize)

def assign_blocks(ordered, n_blocks):
    n = len(ordered)
    base, extra = divmod(n, n_blocks)
    blocks, idx = [], 0
    for i in range(n_blocks):
        size = base + (1 if i < extra else 0)
        blocks.append(ordered[idx:idx+size])
        idx += size
    return blocks

def round_up_5(dt):
    """次の5分刻み (00/05/10/.../55分) に切り上げる。"""
    if dt.second > 0:
        dt = dt.replace(second=0) + timedelta(minutes=1)
    rem = dt.minute % 5
    return dt if rem == 0 else dt + timedelta(minutes=(5 - rem))

MIN_BREAK_MINUTES = 13

def calc_schedule(blocks, durations, start_time_str):
    current = datetime.strptime(start_time_str, "%H:%M")
    schedule = []
    for block_idx, block in enumerate(blocks):
        if block_idx > 0:
            break_start = current
            # 休憩は最低13分確保しつつ、次ブロックの開始は5分刻みに切り上げる
            # （切り上げ幅は最大4分のため、実際の休憩時間は13〜17分に収まる）
            break_end = round_up_5(break_start + timedelta(minutes=MIN_BREAK_MINUTES))
            schedule.append({"type":"break","block":block_idx+1,
                "start":break_start,"end":break_end,
                "duration_sec":int((break_end-break_start).total_seconds())})
            current = break_end
        for name in block:
            dur = durations.get(name, 0)
            end = current + timedelta(seconds=dur)
            schedule.append({"type":"team","block":block_idx+1,"name":name,
                "start":current,"end":end,"duration_sec":dur})
            current = end + timedelta(seconds=30)
    return schedule

def real_gap(schedule, a, b):
    times = {e["name"]:(e["start"],e["end"]) for e in schedule if e["type"]=="team"}
    if a not in times or b not in times: return 0
    sa, ea = times[a]; sb, eb = times[b]
    return int((sb - ea).total_seconds()) if sa <= sb else int((sa - eb).total_seconds())

def check_violations(ordered, durations, conflicts, schedule=None):
    """掛け持ち間隔が最低ライン(40分)を下回っている(=違反)ペアを抽出する。"""
    out = []
    for pair in conflicts:
        a, b = list(pair)
        if a in ordered and b in ordered:
            g = real_gap(schedule, a, b) if schedule else elapsed_gap(ordered, a, b, durations)
            if g < MIN_SHARED_GAP_SEC:
                out.append({"team_a":a,"team_b":b,"gap_sec":g,"gap_str":str(timedelta(seconds=g))})
    return out

def check_shared_notes(ordered, durations, conflicts, schedule=None):
    """掛け持ち間隔は最低ライン(40分)は満たすが、目安(1時間)には届いていないペアを抽出する。"""
    out = []
    for pair in conflicts:
        a, b = list(pair)
        if a in ordered and b in ordered:
            g = real_gap(schedule, a, b) if schedule else elapsed_gap(ordered, a, b, durations)
            if MIN_SHARED_GAP_SEC <= g < IDEAL_SHARED_GAP_SEC:
                out.append({"team_a":a,"team_b":b,"gap_sec":g,"gap_str":str(timedelta(seconds=g))})
    return out

def check_similarity_violations(ordered, pairs):
    """曲/アーティストが同じチーム同士が2組未満しか空いていないペアを抽出する。"""
    out = []
    for pair in pairs:
        a, b = list(pair)
        if a in ordered and b in ordered:
            gap = position_gap(ordered, a, b)
            if gap < MIN_SIMILARITY_GAP:
                out.append({"team_a":a,"team_b":b,"gap":gap})
    return out

def build_remarks(ordered, prefs, violations, shared_notes, song_violations, artist_violations,
                   schedule=None, raw_pref_text=None):
    """文句なしにクリアしていないチームすべてに備考メッセージを組み立てる。
    severity: 'warn' = ルール違反（要対応）, 'note' = 目安未達だが許容範囲（参考情報）"""
    remarks = {name: [] for name in ordered}
    raw_pref_text = raw_pref_text or {}
    actual_starts = {}
    if schedule:
        for e in schedule:
            if e.get("type") == "team":
                actual_starts[e["name"]] = e["start"]

    def add(name, text, severity):
        remarks[name].append((text, severity))

    for v in violations:
        a, b, gap_str = v["team_a"], v["team_b"], v["gap_str"]
        add(a, "掛け持ち間隔不足: {}との間隔が{}（目安:1時間、最低:40分）".format(b, gap_str), "warn")
        add(b, "掛け持ち間隔不足: {}との間隔が{}（目安:1時間、最低:40分）".format(a, gap_str), "warn")
    for v in shared_notes:
        a, b, gap_str = v["team_a"], v["team_b"], v["gap_str"]
        add(a, "掛け持ち間隔が目安未達: {}との間隔が{}（最低40分は確保、目安1時間には未達）".format(b, gap_str), "note")
        add(b, "掛け持ち間隔が目安未達: {}との間隔が{}（最低40分は確保、目安1時間には未達）".format(a, gap_str), "note")
    for v in song_violations:
        a, b, gap = v["team_a"], v["team_b"], v["gap"]
        add(a, "曲の連続: {}との間が{}組（最低2組空け）".format(b, gap), "warn")
        add(b, "曲の連続: {}との間が{}組（最低2組空け）".format(a, gap), "warn")
    for v in artist_violations:
        a, b, gap = v["team_a"], v["team_b"], v["gap"]
        add(a, "アーティストの連続: {}との間が{}組（最低2組空け）".format(b, gap), "warn")
        add(b, "アーティストの連続: {}との間が{}組（最低2組空け）".format(a, gap), "warn")
    n = len(ordered)
    for i, name in enumerate(ordered):
        pref = prefs.get(name)
        if not pref:
            continue
        t = pref["type"]
        raw = str(raw_pref_text.get(name, "") or "").strip()
        prefix = "希望時間帯未達（原文:「{}」）: ".format(raw) if raw else "希望時間帯未達: "
        if t in ("first", "second"):
            actual = "first" if i < n / 2 else "second"
            if actual != t:
                got = "前半" if actual == "first" else "後半"
                add(name, "{}実際は{}に配置".format(prefix, got), "note")
        elif t == "early" and i >= n / 2:
            add(name, "{}出演順{}/{}（後半寄り）".format(prefix, i + 1, n), "note")
        elif t == "late" and i < n / 2:
            add(name, "{}出演順{}/{}（前半寄り）".format(prefix, i + 1, n), "note")
        elif t in ("before", "after"):
            actual_start = actual_starts.get(name)
            if actual_start is None:
                continue
            try:
                deadline = datetime.strptime(pref["time"], "%H:%M").replace(
                    year=actual_start.year, month=actual_start.month, day=actual_start.day)
            except (ValueError, KeyError):
                continue
            if t == "before" and actual_start > deadline:
                miss = str(actual_start - deadline)
                add(name, "{}実際は{}開始（希望より{}超過）".format(prefix, actual_start.strftime("%H:%M"), miss), "note")
            elif t == "after" and actual_start < deadline:
                miss = str(deadline - actual_start)
                add(name, "{}実際は{}開始（希望より{}不足）".format(prefix, actual_start.strftime("%H:%M"), miss), "note")
    return remarks

# ── Excel Output ─────────────────────────────────────────────────────────────

HDR_FILL    = PatternFill("solid", start_color="1F3864")
HDR_FONT    = Font(color="FFFFFF", bold=True, name="Arial")
BLKHDR_FILL = PatternFill("solid", start_color="2E5FA3")
WARN_FILL   = PatternFill("solid", start_color="FFF2CC")
NOTE_FILL   = PatternFill("solid", start_color="E8F4FD")
ALT_FILL    = PatternFill("solid", start_color="F3F3F3")
THIN        = Side(style="thin", color="CCCCCC")

BASE_HEADERS    = ["出演順","ブロック内番号","出演開始時間","チーム名","受付番号","曲名","アーティスト名","出演時間","入りはけ時間"]
BASE_COL_WIDTHS = [9, 13, 16, 18, 12, 22, 18, 11, 13]

def bdr():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def fmt_time(dt): return dt.strftime("%H:%M:%S")

def fmt_dur(sec):
    m, s = divmod(int(sec), 60)
    return "{}:{:02d}".format(m, s)

def apply_hdr(ws, row_idx, fill, font, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row_idx, col)
        c.value = h; c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr()

def write_setlist(ws, schedule, df, durations, remarks):
    team_lookup = df.set_index("name").to_dict("index")
    # 一つでも 'warn'（要対応の違反）があれば黄色、'note'（許容範囲の参考情報）のみなら水色で強調
    warn_teams = {name for name, msgs in remarks.items() if any(sev == "warn" for _, sev in msgs)}
    note_teams = {name for name, msgs in remarks.items() if msgs and name not in warn_teams}

    # 掛け持ちチームの最大数を算出してヘッダーを動的に構築
    max_shared = max((len(info.get("shared_teams", [])) for info in team_lookup.values()), default=0)
    if max_shared == 0:
        shared_headers = ["掛け持ちチーム"]
        shared_widths  = [22]
    else:
        shared_headers = ["掛け持ちチーム{}".format(i+1) for i in range(max_shared)]
        shared_widths  = [18] * max_shared
    headers    = BASE_HEADERS + shared_headers + ["備考"]
    col_widths = BASE_COL_WIDTHS + shared_widths + [40]

    # Row 1: メインヘッダー
    apply_hdr(ws, 1, HDR_FILL, HDR_FONT, headers)

    team_num = 0
    current_block = None
    block_team_num = {}

    for entry in schedule:
        if entry["type"] == "break":
            continue  # 休憩行は出力しない

        block_idx = entry["block"]
        if block_idx != current_block:
            current_block = block_idx
            block_team_num[block_idx] = 0
            # Block 2 以降の先頭にブロックヘッダー行を挿入
            if block_idx >= 2:
                apply_hdr(ws, ws.max_row + 1, BLKHDR_FILL,
                          Font(color="FFFFFF", bold=True, name="Arial"), headers)

        team_num += 1
        block_team_num[block_idx] += 1
        name = entry["name"]
        info = team_lookup.get(name, {})
        block_label = "{}-{}".format(block_idx, block_team_num[block_idx])

        shared_list = info.get("shared_teams", [])
        # 列数に合わせてパディング（掛け持ちなしなら空文字）
        if max_shared == 0:
            shared_cells = [""]
        else:
            shared_cells = shared_list + [""] * (max_shared - len(shared_list))

        remark_text = "; ".join(text for text, _ in remarks.get(name, []))
        row = [team_num, block_label, fmt_time(entry["start"]),
               name, info.get("entry_id",""), info.get("song",""), info.get("artist",""),
               fmt_dur(entry["duration_sec"]), "0:30"] + shared_cells + [remark_text]
        ws.append(row)
        ri = ws.max_row
        if name in warn_teams: fill = WARN_FILL
        elif name in note_teams: fill = NOTE_FILL
        else: fill = ALT_FILL if team_num % 2 == 0 else None
        for col in range(1, len(headers) + 1):
            c = ws.cell(ri, col)
            if fill: c.fill = fill
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="left" if col in (4, len(headers)) else "center", vertical="center")
            c.border = bdr()

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 18

def write_check(ws, ordered, durations, conflicts, schedule):
    hdrs = ["チームA","チームB","実際の間隔","目安/最低","判定"]
    ws.append(hdrs)
    for col in range(1, len(hdrs)+1):
        c = ws.cell(1, col)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center"); c.border = bdr()
    for pair in conflicts:
        a, b = list(pair)
        if a in ordered and b in ordered:
            g = real_gap(schedule, a, b)
            if g >= IDEAL_SHARED_GAP_SEC:
                judge = "OK"
            elif g >= MIN_SHARED_GAP_SEC:
                judge = "許容(40分以上)"
            else:
                judge = "要確認"
            row = [a, b, str(timedelta(seconds=g)), "1時間 / 40分", judge]
            ws.append(row)
            ri = ws.max_row
            for col in range(1, len(hdrs)+1):
                c = ws.cell(ri, col)
                if judge == "要確認": c.fill = WARN_FILL
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(horizontal="center"); c.border = bdr()
    for i in range(1, len(hdrs)+1):
        ws.column_dimensions[get_column_letter(i)].width = 18

def write_similarity_check(ws, ordered, song_pairs, artist_pairs):
    hdrs = ["種別","チームA","チームB","間に挟まる組数","判定"]
    ws.append(hdrs)
    for col in range(1, len(hdrs)+1):
        c = ws.cell(1, col)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center"); c.border = bdr()
    for label, pairs in (("曲", song_pairs), ("アーティスト", artist_pairs)):
        for pair in pairs:
            a, b = list(pair)
            if a in ordered and b in ordered:
                gap = position_gap(ordered, a, b)
                ok = gap >= MIN_SIMILARITY_GAP
                row = [label, a, b, gap, "OK" if ok else "要確認"]
                ws.append(row)
                ri = ws.max_row
                for col in range(1, len(hdrs)+1):
                    c = ws.cell(ri, col)
                    if not ok: c.fill = WARN_FILL
                    c.font = Font(name="Arial", size=10)
                    c.alignment = Alignment(horizontal="center"); c.border = bdr()
    for i in range(1, len(hdrs)+1):
        ws.column_dimensions[get_column_letter(i)].width = 18

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True)
    p.add_argument("--start-time", required=True)
    p.add_argument("--output", required=True)
    p.add_argument("--n-blocks", type=int, default=0)
    p.add_argument("--col-id",     default="受付番号")
    p.add_argument("--col-team",   default="チーム名")
    p.add_argument("--col-song",   default="曲名")
    p.add_argument("--col-artist", default="アーティスト名")
    p.add_argument("--col-shared", default="掛け持ちチーム")
    p.add_argument("--col-pref",   default="希望時間帯")
    p.add_argument("--duration-overrides", default="")
    p.add_argument("--pref-overrides", default="",
                   help='希望時間帯の構造化上書き。例: \'{"TEAM_A":{"type":"early"},"TEAM_B":{"type":"before","time":"15:00"}}\'')
    p.add_argument("--seed", type=int, default=None,
                   help="出演順タイブレークをランダム化する乱数シード。同じ入力に対し--seedを変えて"
                        "複数回実行し、末尾に出力される QUALITY_SUMMARY を比較することで、"
                        "貪欲法が局所最適に陥った場合の代替案を探索できる。省略時は従来通り決定的。")
    args = p.parse_args()
    if args.seed is not None:
        random.seed(args.seed)

    path = Path(args.input)
    df = pd.read_excel(path) if path.suffix in (".xlsx",".xls") else pd.read_csv(path)
    df = strip_trailing_notes(df)
    col_map = {args.col_id:"entry_id", args.col_team:"name", args.col_song:"song", args.col_artist:"artist",
               args.col_shared:"shared_teams_raw", args.col_pref:"preferred_time"}
    df = df.rename(columns={k:v for k,v in col_map.items() if k in df.columns})
    if "shared_teams_raw" in df.columns:
        df["shared_teams"] = df["shared_teams_raw"].apply(parse_shared)
    else:
        df["shared_teams"] = [[] for _ in range(len(df))]
    if "entry_id" not in df.columns:
        df["entry_id"] = ""
    teams = df.to_dict("records")

    overrides = json.loads(args.duration_overrides) if args.duration_overrides else {}
    client_id, client_secret = load_spotify_credentials()
    durations = {}; missing = []

    if client_id and client_secret:
        try:
            token = get_spotify_token(client_id, client_secret)
            for t in teams:
                if t["name"] in overrides:
                    durations[t["name"]] = int(overrides[t["name"]]); continue
                sec = search_track_duration(token, t["song"], t["artist"])
                if sec: durations[t["name"]] = sec
                else: missing.append(t); durations[t["name"]] = overrides.get(t["name"],0)
        except Exception as e:
            sys.stderr.write("[Spotify Error] {}\n".format(e))
            missing = [t for t in teams if t["name"] not in overrides]
            for t in teams: durations[t["name"]] = overrides.get(t["name"],0)
    else:
        sys.stderr.write("Spotify認証情報なし\n")
        missing = [t for t in teams if t["name"] not in overrides]
        for t in teams: durations[t["name"]] = overrides.get(t["name"],0)

    if missing:
        print("[手動入力が必要な曲]")
        for t in missing:
            print("  {}: {} / {}".format(t["name"], t["song"], t["artist"]))

    conflicts = build_conflicts(teams)
    song_pairs = build_similarity_conflicts(teams, "song")
    artist_pairs = build_similarity_conflicts(teams, "artist")
    pref_overrides = json.loads(args.pref_overrides) if args.pref_overrides else {}
    n_blocks = args.n_blocks if args.n_blocks > 0 else (5 if len(teams) > 15 else 4)
    start_dt = datetime.strptime(args.start_time, "%H:%M")
    ordered = greedy_schedule(teams, durations, conflicts, n_blocks=n_blocks, break_sec=MIN_BREAK_MINUTES * 60,
                               song_pairs=song_pairs, artist_pairs=artist_pairs,
                               pref_overrides=pref_overrides, start_dt=start_dt,
                               randomize=(args.seed is not None))
    blocks = assign_blocks(ordered, n_blocks)
    schedule = calc_schedule(blocks, durations, args.start_time)
    violations = check_violations(ordered, durations, conflicts, schedule=schedule)
    shared_notes = check_shared_notes(ordered, durations, conflicts, schedule=schedule)
    song_violations = check_similarity_violations(ordered, song_pairs)
    artist_violations = check_similarity_violations(ordered, artist_pairs)
    prefs = {t["name"]: normalize_pref(pref_overrides.get(t["name"]), t.get("preferred_time")) for t in teams}
    raw_pref_text = {t["name"]: t.get("preferred_time") for t in teams}
    remarks = build_remarks(ordered, prefs, violations, shared_notes, song_violations, artist_violations,
                             schedule=schedule, raw_pref_text=raw_pref_text)

    wb = Workbook()
    ws1 = wb.active; ws1.title = "セットリスト"
    write_setlist(ws1, schedule, df, durations, remarks)
    if conflicts:
        ws2 = wb.create_sheet("掛け持ちチェック")
        write_check(ws2, ordered, durations, conflicts, schedule)
    if song_pairs or artist_pairs:
        ws3 = wb.create_sheet("曲・アーティスト重複チェック")
        write_similarity_check(ws3, ordered, song_pairs, artist_pairs)
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)

    team_entries = [e for e in schedule if e["type"]=="team"]
    print("SETLIST COMPLETE: {}".format(args.output))
    print("  Teams:{} Blocks:{}".format(len(teams), n_blocks))
    if team_entries:
        print("  End time: {}".format(fmt_time(team_entries[-1]["end"])))
    if violations:
        print("WARNING: {} conflict(s):".format(len(violations)))
        for v in violations:
            print("  {} <-> {}: {}".format(v["team_a"], v["team_b"], v["gap_str"]))
    if song_violations:
        print("WARNING: {} song repeat(s) within 2 slots".format(len(song_violations)))
    if artist_violations:
        print("WARNING: {} artist repeat(s) within 2 slots".format(len(artist_violations)))

    # --seed を変えて複数回実行し比較するための定量サマリー。
    # warn: 要対応の違反件数(少ないほど良い) / note: 許容範囲の参考件数(少ないほど良い)
    # min_shared_gap_sec: 掛け持ちペアの中で最も短い間隔(秒、大きいほど良い＝偏りが少ない)
    warn_count = sum(1 for msgs in remarks.values() if any(sev == "warn" for _, sev in msgs))
    note_count = sum(1 for msgs in remarks.values() if msgs and not any(sev == "warn" for _, sev in msgs))
    if conflicts:
        min_shared_gap = min(real_gap(schedule, *list(pair)) for pair in conflicts)
    else:
        min_shared_gap = None
    print("QUALITY_SUMMARY: warn={} note={} min_shared_gap_sec={}".format(
        warn_count, note_count, min_shared_gap if min_shared_gap is not None else "NA"))
    if missing:
        print("WARNING: {} song(s) not found".format(len(missing)))

if __name__ == "__main__":
    main()
